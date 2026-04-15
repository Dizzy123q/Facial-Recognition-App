import sys
import time
from openpyxl import load_workbook

from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *

class MyDialog(QDialog):
    def __init__(self, nume_persoana):
        super().__init__()

        self.nume = nume_persoana

        Layout_vertical = QVBoxLayout()
        Layout_orizontal = QHBoxLayout()

        self.setWindowTitle("Person identified!")
        self.setWindowIcon(QIcon('face-scan.png'))

        mesaj = QLabel(f"Person identified:  {self.nume.split('.')[-3]}")
        mesaj2 = QLabel(f"Confidence:  {self.nume.split(' ')[-1]}")
        Layout_vertical.addWidget(mesaj)
        Layout_vertical.addWidget(mesaj2)

        # Create three buttons
        self.buton_adauga_in_istoric = QPushButton("Add to history", self)
        self.buton_adauga_in_istoric.clicked.connect(self.adauga_in_istoric)
        Layout_orizontal.addWidget(self.buton_adauga_in_istoric)

        self.buton_scanare = QPushButton("Continue scanning", self)
        self.buton_scanare.clicked.connect(self.continua_scanarea)
        # Add the buttons to a layout


        Layout_orizontal.addWidget(self.buton_scanare)
        Layout_vertical.addLayout(Layout_orizontal)

        # Set the layout for the dialog
        self.setLayout(Layout_vertical)

    def adauga_in_istoric(self):
        ora_curenta = time.localtime()
        data_si_ora = time.strftime("%Y-%m-%d %H:%M:%S", ora_curenta)
        print(data_si_ora)

        self.valoare1 = self.nume.split(".")[-3]
        self.valoare2 = data_si_ora
        print(self.valoare1 + " " + self.valoare2)

        workbook = load_workbook('logs.xlsx')
        worksheet = workbook.active

        next_row = worksheet.max_row + 1
        worksheet.cell(row=next_row, column=1).value = self.valoare1
        worksheet.cell(row=next_row, column=2).value = self.valoare2
        workbook.save('logs.xlsx')

    def continua_scanarea(self):
        self.close()

def main():
    aplicatie = QApplication([])
    dialog = MyDialog("Test 0.0%")
    dialog.show()
    sys.exit(aplicatie.exec_())

if __name__ == '__main__':
    main()


